"""Admin API endpoints for user management.

All endpoints require HR role access.
"""
import io
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse
from sqlalchemy import select
from sqlalchemy.orm import Session
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

from app import models
from app.auth.rbac import RequireHR
from app.core.errors import AppError
from app.db.session import get_session
from app.auth.password import hash_password, validate_password_strength
from app.models.domain import User, UserRole
from app.schemas.base import MessageResponse, UserCreate, UserRead, UserUpdate

router = APIRouter(prefix="/admin", tags=["admin"])


@router.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(
    user_data: UserCreate,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> UserRead:
    """
    Create a new user.
    
    Requires HR role.
    
    - Email must be unique
    - Password must meet strength requirements
    - Role must be a valid UserRole enum value (EMPLOYEE, TEAM_LEAD, MANAGER, HR)
    - HR can assign any role including HR (admin) role - no restrictions
    - Status defaults to ACTIVE if not provided
    - Team must exist if team_id is provided
    """
    from uuid import uuid4
    
    # Validate password strength
    is_valid, error_message = validate_password_strength(user_data.password)
    if not is_valid:
        raise AppError(
            error_message or "Invalid password",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if email already exists
    existing_user = db.scalar(select(User).where(User.email == user_data.email))
    if existing_user:
        raise AppError(
            "Email already in use",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate role
    try:
        role_enum = UserRole[user_data.role.upper()]
    except KeyError:
        raise AppError(
            f"Invalid role. Must be one of: {[r.name for r in UserRole]}",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate status
    status_value = (user_data.status or "ACTIVE").upper()
    if status_value not in ("ACTIVE", "INACTIVE"):
        raise AppError(
            "Invalid status. Must be ACTIVE or INACTIVE",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    # Validate team_id if provided
    if user_data.team_id:
        team = db.get(models.Team, user_data.team_id)
        if not team:
            raise AppError(
                "Team not found",
                status_code=status.HTTP_400_BAD_REQUEST
            )
    
    try:
        # Hash password
        password_hash = hash_password(user_data.password)
        
        # Create user
        user = User(
            id=uuid4(),
            name=user_data.name,
            email=user_data.email,
            password_hash=password_hash,
            role=role_enum,
            team_id=user_data.team_id,
            status=status_value,
        )
        
        db.add(user)
        db.commit()
        db.refresh(user)
        
        # Enrich user with team name
        user_dict = UserRead.model_validate(user).model_dump()
        if user.team:
            user_dict['team_name'] = user.team.name
        
        return UserRead.model_validate(user_dict)
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create user: {str(e)}"
        )


@router.get("/users", response_model=List[UserRead])
def list_users(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    role_filter: Optional[str] = Query(None, description="Filter by role (EMPLOYEE, TEAM_LEAD, MANAGER, HR)"),
    status_filter: Optional[str] = Query(None, description="Filter by status (ACTIVE, INACTIVE)"),
    team_id: Optional[UUID] = Query(None, description="Filter by team ID"),
    search: Optional[str] = Query(None, description="Search by name or email"),
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> List[UserRead]:
    """
    List all users with optional filtering.
    
    Requires HR role.
    """
    stmt = select(User)
    
    # Apply filters
    if role_filter:
        try:
            role_enum = UserRole[role_filter.upper()]
            stmt = stmt.where(User.role == role_enum)
        except KeyError:
            raise AppError(
                f"Invalid role. Must be one of: {[r.name for r in UserRole]}",
                status_code=status.HTTP_400_BAD_REQUEST
            )
    
    if status_filter:
        stmt = stmt.where(User.status == status_filter.upper())
    
    if team_id:
        stmt = stmt.where(User.team_id == team_id)
    
    if search:
        search_pattern = f"%{search}%"
        stmt = stmt.where(
            (User.name.ilike(search_pattern)) | (User.email.ilike(search_pattern))
        )
    
    stmt = stmt.order_by(User.created_at.desc()).offset(skip).limit(limit)
    # Eager load team relationship to avoid N+1 queries
    from sqlalchemy.orm import joinedload
    stmt = stmt.options(joinedload(User.team))
    users = db.scalars(stmt).unique().all()
    
    # Enrich users with team names
    result = []
    for user in users:
        user_dict = UserRead.model_validate(user).model_dump()
        if user.team:
            user_dict['team_name'] = user.team.name
        result.append(UserRead.model_validate(user_dict))
    
    return result


@router.get("/users/{user_id}", response_model=UserRead)
def get_user(
    user_id: UUID,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> UserRead:
    """
    Get a specific user by ID.
    
    Requires HR role.
    """
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Enrich user with team name
    user_dict = UserRead.model_validate(user).model_dump()
    if user.team:
        user_dict['team_name'] = user.team.name
    
    return UserRead.model_validate(user_dict)


@router.patch("/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: UUID,
    user_update: UserUpdate,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> UserRead:
    """
    Update a user's information (role, status, team, name, email).
    
    Requires HR role.
    
    - Cannot update password through this endpoint (use password reset flow)
    - Email uniqueness is validated if email is being updated
    - Role must be a valid UserRole enum value (EMPLOYEE, TEAM_LEAD, MANAGER, HR)
    - HR can assign or remove any role including HR (admin) role - no restrictions
    - Status must be ACTIVE or INACTIVE
    """
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    update_data = user_update.model_dump(exclude_unset=True)
    
    # Validate email uniqueness if email is being updated
    if "email" in update_data and update_data["email"] != user.email:
        existing_user = db.scalar(select(User).where(User.email == update_data["email"]))
        if existing_user and existing_user.id != user_id:
            raise AppError(
                "Email already in use by another user",
                status_code=status.HTTP_400_BAD_REQUEST
            )
        user.email = update_data["email"]
    
    # Update name if provided
    if "name" in update_data:
        user.name = update_data["name"]
    
    # Update role if provided
    if "role" in update_data:
        try:
            user.role = UserRole[update_data["role"].upper()]
        except KeyError:
            raise AppError(
                f"Invalid role. Must be one of: {[r.name for r in UserRole]}",
                status_code=status.HTTP_400_BAD_REQUEST
            )
    
    # Update status if provided
    if "status" in update_data:
        status_value = update_data["status"].upper()
        if status_value not in ("ACTIVE", "INACTIVE"):
            raise AppError(
                "Invalid status. Must be ACTIVE or INACTIVE",
                status_code=status.HTTP_400_BAD_REQUEST
            )
        user.status = status_value
    
    # Update team_id if provided
    if "team_id" in update_data:
        team_id = update_data["team_id"]
        if team_id is not None:
            # Validate team exists
            team = db.get(models.Team, team_id)
            if not team:
                raise AppError(
                    "Team not found",
                    status_code=status.HTTP_400_BAD_REQUEST
                )
        user.team_id = team_id
    
    try:
        db.commit()
        db.refresh(user)
        
        # Enrich user with team name
        user_dict = UserRead.model_validate(user).model_dump()
        if user.team:
            user_dict['team_name'] = user.team.name
        
        return UserRead.model_validate(user_dict)
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update user: {str(e)}"
        )


@router.delete("/users/{user_id}", response_model=MessageResponse)
def delete_user(
    user_id: UUID,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> MessageResponse:
    """
    Delete a user (soft delete by setting status to INACTIVE).
    
    Requires HR role.
    
    Note: This performs a soft delete by setting the user status to INACTIVE
    rather than actually deleting the record, to preserve data integrity.
    
    To permanently delete, contact database administrator.
    """
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        raise AppError(
            "Cannot delete your own account",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    # Soft delete: set status to INACTIVE
    user.status = "INACTIVE"
    
    try:
        db.commit()
        return MessageResponse(message=f"User {user.email} has been deactivated successfully")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete user: {str(e)}"
        )


@router.post("/users/{user_id}/activate", response_model=MessageResponse)
def activate_user(
    user_id: UUID,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> MessageResponse:
    """
    Activate a user (set status to ACTIVE).
    
    Requires HR role.
    
    Enables a user account that was previously deactivated.
    """
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    user.status = "ACTIVE"
    
    try:
        db.commit()
        return MessageResponse(message=f"User {user.email} has been activated successfully")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to activate user: {str(e)}"
        )


@router.post("/users/{user_id}/deactivate", response_model=MessageResponse)
def deactivate_user(
    user_id: UUID,
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> MessageResponse:
    """
    Deactivate a user (set status to INACTIVE).
    
    Requires HR role.
    
    Disables a user account. Deactivated users cannot log in.
    This is equivalent to soft delete and can be reversed with activate endpoint.
    """
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Prevent deactivating yourself
    if user.id == current_user.id:
        raise AppError(
            "Cannot deactivate your own account",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    user.status = "INACTIVE"
    
    try:
        db.commit()
        return MessageResponse(message=f"User {user.email} has been deactivated successfully")
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to deactivate user: {str(e)}"
        )


@router.post("/users/bulk", response_class=JSONResponse, status_code=status.HTTP_200_OK)
async def create_users_bulk(
    file: UploadFile = File(...),
    current_user: User = Depends(RequireHR),
    db: Session = Depends(get_session),
) -> dict:
    """
    Create multiple users from an Excel file.
    
    Requires HR role.
    
    Excel file format:
    - First row should contain headers: Name, Email, Password, Role, Status (optional), Department (optional)
    - Required columns: Name, Email, Password, Role
    - Optional columns: Status, Department (must match existing department name)
    - Status defaults to ACTIVE if not provided
    - Department is matched by name (case-insensitive)
    
    Returns:
    - List of successfully created users
    - List of failed users with error messages
    - Summary statistics
    """
    from uuid import uuid4
    import re
    
    # Check if openpyxl is available
    if not OPENPYXL_AVAILABLE:
        raise AppError(
            "openpyxl library is not installed. Please install it: pip install openpyxl",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    # Validate file type
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise AppError(
            "Invalid file type. Please upload an Excel file (.xlsx or .xls)",
            status_code=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Read file content (await for async)
        file_content = await file.read()
        if not file_content:
            raise AppError("File is empty", status_code=status.HTTP_400_BAD_REQUEST)
        
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise AppError("Excel file is empty or invalid", status_code=status.HTTP_400_BAD_REQUEST)
        
        # Normalize headers (lowercase, strip whitespace)
        headers = [str(h).strip().lower() if h else "" for h in headers]
        
        # Find column indices
        name_idx = None
        email_idx = None
        password_idx = None
        role_idx = None
        status_idx = None
        department_idx = None
        
        for idx, header in enumerate(headers):
            if 'name' in header:
                name_idx = idx
            elif 'email' in header:
                email_idx = idx
            elif 'password' in header:
                password_idx = idx
            elif 'role' in header:
                role_idx = idx
            elif 'status' in header:
                status_idx = idx
            elif 'department' in header:
                department_idx = idx
        
        # Validate required columns
        missing_cols = []
        if name_idx is None:
            missing_cols.append("Name")
        if email_idx is None:
            missing_cols.append("Email")
        if password_idx is None:
            missing_cols.append("Password")
        if role_idx is None:
            missing_cols.append("Role")
        
        if missing_cols:
            raise AppError(
                f"Missing required columns: {', '.join(missing_cols)}",
                status_code=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all existing emails to check for duplicates
        existing_users = db.scalars(select(User)).all()
        existing_emails = {user.email.lower() for user in existing_users}
        
        # Get all teams for department matching
        teams = {team.name.upper(): team.id for team in db.scalars(select(models.Team)).all()}
        
        # Process rows
        results = {
            "success": [],
            "failed": [],
            "summary": {
                "total": 0,
                "created": 0,
                "failed": 0,
                "skipped": 0
            }
        }
        
        # Process each row (skip header row)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            results["summary"]["total"] += 1
            row_data = {}
            errors = []
            
            # Extract values
            name = str(row[name_idx]).strip() if row[name_idx] else ""
            email = str(row[email_idx]).strip().lower() if row[email_idx] else ""
            password = str(row[password_idx]).strip() if row[password_idx] else ""
            role = str(row[role_idx]).strip().upper() if row[role_idx] else ""
            user_status = str(row[status_idx]).strip().upper() if status_idx is not None and row[status_idx] else "ACTIVE"
            department_name = str(row[department_idx]).strip() if department_idx is not None and row[department_idx] else ""
            
            # Validate name
            if not name or len(name) > 255:
                errors.append("Name is required and must be 255 characters or less")
            
            # Validate email
            if not email:
                errors.append("Email is required")
            elif not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
                errors.append("Invalid email format")
            elif email in existing_emails:
                # Mark as skipped (duplicate email)
                results["failed"].append({
                    "row": row_num,
                    "name": name or "N/A",
                    "email": email,
                    "errors": ["Email already exists in system"]
                })
                results["summary"]["skipped"] += 1
                continue
            
            # Validate password
            if not password:
                errors.append("Password is required")
            else:
                is_valid, error_message = validate_password_strength(password)
                if not is_valid:
                    errors.append(f"Password: {error_message}")
            
            # Validate role
            if not role:
                errors.append("Role is required")
            else:
                try:
                    UserRole[role]
                except KeyError:
                    errors.append(f"Invalid role. Must be one of: {[r.name for r in UserRole]}")
            
            # Validate status
            if user_status and user_status not in ("ACTIVE", "INACTIVE"):
                errors.append("Status must be ACTIVE or INACTIVE")
            
            # Validate department
            team_id = None
            if department_name:
                department_upper = department_name.upper()
                if department_upper in teams:
                    team_id = teams[department_upper]
                else:
                    errors.append(f"Department '{department_name}' not found")
            
            # If there are errors, add to failed list
            if errors:
                results["failed"].append({
                    "row": row_num,
                    "name": name or "N/A",
                    "email": email or "N/A",
                    "errors": errors
                })
                results["summary"]["failed"] += 1
                continue
            
            # Check if email already exists in this batch (duplicate in file)
            if email in [r.get("email", "").lower() for r in results["success"]]:
                results["failed"].append({
                    "row": row_num,
                    "name": name,
                    "email": email,
                    "errors": ["Duplicate email in this file"]
                })
                results["summary"]["failed"] += 1
                continue
            
            # Try to create user
            try:
                # Hash password
                password_hash = hash_password(password)
                
                # Create user
                user = User(
                    id=uuid4(),
                    name=name,
                    email=email,
                    password_hash=password_hash,
                    role=UserRole[role],
                    team_id=team_id,
                    status=user_status or "ACTIVE",
                )
                
                db.add(user)
                db.flush()  # Flush to check for unique constraint violations
                
                # Add to existing emails to prevent duplicates in same batch
                existing_emails.add(email.lower())
                
                # Enrich with team name
                user_dict = {
                    "id": str(user.id),
                    "name": user.name,
                    "email": user.email,
                    "role": user.role.value,
                    "status": user.status,
                    "row": row_num
                }
                if user.team:
                    user_dict["team_name"] = user.team.name
                
                results["success"].append(user_dict)
                results["summary"]["created"] += 1
                
            except Exception as e:
                db.rollback()
                error_msg = str(e)
                if "unique constraint" in error_msg.lower() or "duplicate" in error_msg.lower():
                    error_msg = "Email already exists"
                    results["summary"]["skipped"] += 1
                else:
                    results["summary"]["failed"] += 1
                
                results["failed"].append({
                    "row": row_num,
                    "name": name,
                    "email": email,
                    "errors": [error_msg]
                })
        
        # Check if any rows were processed
        if results["summary"]["total"] == 0:
            raise AppError(
                "No data rows found in Excel file. Please ensure the file has data rows after the header.",
                status_code=status.HTTP_400_BAD_REQUEST
            )
        
        # Commit all successful creations
        try:
            if results["summary"]["created"] > 0:
                db.commit()
            else:
                db.rollback()  # No point committing if nothing was created
        except Exception as e:
            db.rollback()
            raise AppError(
                f"Failed to create users: {str(e)}",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        return results
        
    except AppError as e:
        # Re-raise AppError as-is (it will be handled by app_error_handler with CORS)
        raise
    except ImportError as e:
        if "openpyxl" in str(e):
            raise AppError(
                "openpyxl library is not installed. Please install it: pip install openpyxl",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        raise AppError(
            f"Import error: {str(e)}",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        import traceback
        error_detail = str(e)
        # Log the full traceback for debugging
        import structlog
        logger = structlog.get_logger()
        logger.error("bulk_user_upload_error", error=error_detail, traceback=traceback.format_exc())
        
        # Use AppError instead of HTTPException to ensure CORS headers are included
        raise AppError(
            f"Failed to process Excel file: {error_detail}",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
